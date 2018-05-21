import requests
from openpyxl import load_workbook
from datetime import date, timedelta,datetime
import csv

class FileGenerator:
	"""FileGenerator used to generate csv file from a set of excel file which having specific data."""
	
	def __init__(self):
		self.csv_1_header = ['Date','BCB_Commercial_Exports_Total','BCB_Commercial_Exports_Advances_on_Contracts','BCB_Commercial_Exports_Payment_Advance','BCB_Commercial_Exports_Others','BCB_Commercial_Imports','BCB_Commercial_Balance','BCB_Financial_Purchases','BCB_Financial_Sales','BCB_Financial_Balance','BCB_Balance']
		self.source_data_path = {"BCB_Commercial_Exports_Total":"C","BCB_Commercial_Exports_Advances_on_Contracts":"D","BCB_Commercial_Exports_Payment_Advance":"E","BCB_Commercial_Exports_Others":"F","BCB_Commercial_Imports":"G","BCB_Commercial_Balance":"H","BCB_Financial_Purchases":"I","BCB_Financial_Sales":"J","BCB_Financial_Balance":"K","BCB_Balance":"L"}
		self.csv_2_header = ["Date","BCB_FX_Position"]
		self.abr_month = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
		self.x_range = 1000

	def download_file(self,url):
		"""Download a csv file. 
            Args
            ----
            url --> the url will point to a csv file.

            Return
            ------
            file_name --> Name of the file downloaded.
        """
		file_name = url.split('/')[-1]
		res = requests.get(url)
		with open(file_name,'wb') as fp:
			fp.write(res.content)
		return file_name

	def create_csv_file(self,data,csv_header,out_file):
		"""Create a csv file with a specific header and name. 
            Args
            ----
            data --> The data need to write on the csv file.
            csv_header --> csv header for the writting.
            out_file --> The desired name of file to write.
        """
		with open(out_file, 'wb') as myfile:
			wr = csv.writer(myfile)
			wr.writerow(csv_header)
			for row in data:
				wr.writerow(row)

	def get_date_type1(self,date,file_name):
		"""Fetch the values from the type 1 files,
            Args
            ----
            date --> The desired date we need to get the data.
            file_name --> Name of file which have the data.

            Return
            ------
            the integer value or None.
        """
		end_limit = 0
		year_limit = 0
		month_limit = 0
		day_limit = 0

		workbook = load_workbook(file_name)
		worksheet = workbook.active
		
		month,day,year = date.split('/')
		for i in range(1,self.x_range):
			if worksheet['A'+str(i)].value == 'Memo:':
				end_limit = i
				break
		if end_limit:
			for iy in range(1,end_limit):
				if worksheet['A'+str(iy)].value == int(year):
					year_limit = iy
					break
			if year_limit:
				for im in range(year_limit,end_limit):
					if worksheet['B'+str(im)].value == self.abr_month[int(month)]:
						month_limit = im
						break

				if month_limit:
					if not isinstance(worksheet['B'+str(month_limit+1)].value,long):
						day_limit = month_limit
					else:
						for idate in range(month_limit,month_limit+30):
							if worksheet['B'+str(idate)].value == int(day):
								day_limit = idate
								break	
						if not day_limit:
							day_limit = month_limit
					return {key:worksheet[val+str(day_limit)].value for key,val in self.source_data_path.items()}
				else:
					print "Month limit not found"
			else:
				print "Year liit not found"
		else:
			"Wrong file format"

	def get_date_type2(self,date,file_name):
		"""Fetch the values from the type 2 files,
            Args
            ----
            date --> The desired date we need to get the data.
            file_name --> Name of file which have the data.

            Return
            ------
            the integer value or None.
        """
		year_limit = 0
		month_limit = 0
		
		workbook = load_workbook(file_name)
		worksheet = workbook.active
		
		month,day,year = date.split('/')
		for iy in range(1,self.x_range):
			if worksheet['A'+str(iy)].value == int(year):
				year_limit = iy
				break
		if year_limit:
			for im in range(year_limit,year_limit+13):
				if worksheet['B'+str(im)].value == self.abr_month[int(month)]:
					month_limit = im
					break

			if month_limit:
				return worksheet["C"+str(month_limit)].value
			else:
				print "Month limit not found"
		else:
			print "Year liit not found"

	def generate_file(self,type,date_ranges,out_file_name):
		"""Generate the final output file.
            Args
            ----
            type --> Type of file, Either 'type1' or 'type2'.
            date_ranges --> It's may be date or range of date in format 'mm/dd/yyyy' or 'mm/dd/yyyy - mm/dd/yyyy'.
            out_file_name --> Desired output file name

     	"""
		if '-' in date_ranges:
			start_date_str,end_date_str = date_ranges.split('-')
		else:
			start_date_str = end_date_str = date_ranges
		
		start_date = datetime.strptime(start_date_str,"%m/%d/%Y")
		end_date = datetime.strptime(end_date_str,"%m/%d/%Y")
		delta = end_date - start_date
		result = []
		if type == 'type1':
			key_to_sort = self.csv_1_header[1:]
			file_name = self.download_file("http://www.bcb.gov.br/pec/Indeco/Ingl/ie5-24i.xlsx")
			for i in range(delta.days + 1):
				date = (start_date + timedelta(days=i)).strftime("%m/%d/%Y")
				data = self.get_date_type1(date,file_name)
				values_to_insert_ = sorted(data.items(),key=lambda x: key_to_sort.index(x[0]))
				values_to_insert = [i[1] for i in values_to_insert_]
				result.append([date]+values_to_insert)
			self.create_csv_file(result,self.csv_1_header,out_file_name)
			print 'type 1 file csv has been generted'

		elif type == 'type2':
			file_name = self.download_file("http://www.bcb.gov.br/pec/Indeco/Ingl/ie5-26i.xlsx")
			for i in range(delta.days + 1):
				date = (start_date + timedelta(days=i)).strftime("%m/%d/%Y")
				data = self.get_date_type2(date,file_name)
				result.append([date,data])

			self.create_csv_file(result,self.csv_2_header,out_file_name)
			print 'type 2 file csv has been generted'

		
if __name__ == '__main__':
	file_gen_obj = FileGenerator()
	file_gen_obj.generate_file('type1','12/1/2017','test_out_1.csv')
	file_gen_obj.generate_file('type2','3/26/2018-3/28/2018','test_out_2.csv')
